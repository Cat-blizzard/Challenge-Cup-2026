"""CLI for MathSolve-Agent."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from .parser import MathSolution, fallback_solution, parse_and_validate, solution_to_json


ID_FIELDS = ("problem_id", "id", "question_id", "qid", "uid", "index")
TEXT_FIELDS = ("problem_text", "problem", "question", "text", "content", "题目")

SAMPLE_PROBLEMS = [
    {
        "problem_id": "001",
        "problem_text": "Find all real roots of x^4 - 5x^2 + 4 = 0.",
    },
    {
        "problem_id": "002",
        "problem_text": (
            "Given f(z) = (z^2 + 1)/(z - i), find the residue at z = i."
        ),
    },
    {
        "problem_id": "003",
        "problem_text": (
            "Maximize 3x + 4y subject to x + 2y <= 8, 3x + y <= 9, "
            "x >= 0, y >= 0."
        ),
    },
]


def run_single_demo(
    model_type: str = "gpt-4o-mini",
    api_key: Optional[str] = None,
    api_base: Optional[str] = None,
    config_path: Optional[str] = None,
    ablation: str = "full",
    official_mode: bool = False,
) -> None:
    from .agent import MathSolverAgent

    agent = MathSolverAgent(
        model_type=model_type,
        api_key=api_key,
        api_base=api_base,
        config_path=config_path,
        ablation=ablation,
        official_mode=official_mode,
    )
    item = SAMPLE_PROBLEMS[0]
    start = time.time()
    solution = agent.solve(item["problem_text"], item["problem_id"])
    elapsed = time.time() - start

    print("=" * 72)
    print(f"Problem [{item['problem_id']}]: {item['problem_text']}")
    print(f"Elapsed: {elapsed:.1f}s")
    print(f"Passed: {solution.verification.passed}")
    print(f"Confidence: {solution.verification.confidence:.2f}")
    print(f"Answer: {solution.answer}")
    print("-" * 72)
    print(solution_to_json(solution))


def run_batch(
    input_path: str,
    output_path: str,
    model_type: str = "gpt-4o-mini",
    api_key: Optional[str] = None,
    api_base: Optional[str] = None,
    limit: Optional[int] = None,
    resume: bool = False,
    results_json_path: Optional[str] = None,
    log_dir: Optional[str] = None,
    summary_path: Optional[str] = None,
    config_path: Optional[str] = None,
    ablation: str = "full",
    official_mode: bool = False,
) -> Dict[str, Any]:
    from .agent import MathSolverAgent

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    results_json = Path(results_json_path) if results_json_path else output.with_suffix(".json")
    logs = Path(log_dir) if log_dir else output.parent / "logs"
    summary = Path(summary_path) if summary_path else output.parent / "run_summary.json"
    logs.mkdir(parents=True, exist_ok=True)

    problems = load_problems(input_path)
    if limit is not None:
        problems = problems[:limit]

    existing_results: Dict[str, Dict[str, Any]] = {}
    if resume and output.exists():
        existing_results = _read_existing_results(output)

    mode = "a" if resume and output.exists() else "w"
    agent = MathSolverAgent(
        model_type=model_type,
        api_key=api_key,
        api_base=api_base,
        config_path=config_path,
        ablation=ablation,
        official_mode=official_mode,
    )

    total = len(problems)
    processed = 0
    skipped = 0
    fallback_count = 0
    started_at = time.time()

    print(f"Loaded {total} problems from {input_path}")
    print(f"Writing JSONL to {output}")
    print(f"Writing per-problem logs to {logs}")

    with output.open(mode, encoding="utf-8") as f_out:
        for index, record in enumerate(problems, start=1):
            pid = record["problem_id"]
            text = record["problem_text"]
            metadata = record.get("raw_metadata", {})

            if resume and pid in existing_results:
                skipped += 1
                print(f"[{index}/{total}] skip {pid} (resume)")
                continue

            print(f"[{index}/{total}] solving {pid} ...")
            item_start = time.time()
            used_exception_fallback = False
            try:
                solution = agent.solve(text, pid, raw_metadata=metadata)
                run_log = dict(agent.last_run_log)
            except Exception as exc:
                used_exception_fallback = True
                solution = fallback_solution(pid, f"{type(exc).__name__}: {exc}")
                run_log = {
                    "problem_id": pid,
                    "raw_problem": text,
                    "raw_metadata": metadata,
                    "api_status": "batch_fallback_exception",
                    "exception": repr(exc),
                    "latency_seconds": round(time.time() - item_start, 3),
                    "final_json": solution.model_dump(mode="json"),
                }

            if (
                used_exception_fallback
                or solution.answer == "unable_to_determine"
                or not solution.verification.passed
            ):
                fallback_count += 1

            line = solution_to_json(solution, indent=None)
            f_out.write(line + "\n")
            f_out.flush()
            _write_problem_log(logs, pid, run_log)
            processed += 1

            elapsed = time.time() - item_start
            print(
                f"  done in {elapsed:.1f}s | passed={solution.verification.passed} "
                f"| conf={solution.verification.confidence:.2f} | answer={solution.answer[:80]}"
            )

    all_results, schema_errors = _load_and_validate_results(output)
    results_json.write_text(
        json.dumps(all_results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    run_summary = {
        "input_path": str(input_path),
        "output_jsonl": str(output),
        "output_json": str(results_json),
        "log_dir": str(logs),
        "total_loaded": total,
        "processed_this_run": processed,
        "skipped_by_resume": skipped,
        "results_in_jsonl": len(all_results),
        "schema_error_count": len(schema_errors),
        "schema_errors": schema_errors[:20],
        "fallback_or_unpassed_count_this_run": fallback_count,
        "config_path": config_path,
        "ablation": ablation,
        "elapsed_seconds": round(time.time() - started_at, 3),
    }
    summary.write_text(json.dumps(run_summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=" * 72)
    print(f"Batch complete. Processed={processed}, skipped={skipped}, results={len(all_results)}")
    print(f"Schema errors={len(schema_errors)}")
    print(f"Final JSON: {results_json}")
    print(f"Summary: {summary}")
    return run_summary


def load_problems(path: str) -> List[Dict[str, Any]]:
    source = Path(path)
    suffix = source.suffix.lower()
    if suffix == ".jsonl":
        rows = _load_jsonl(source)
    elif suffix == ".json":
        rows = _load_json(source)
    elif suffix == ".csv":
        rows = _load_csv(source)
    elif suffix in {".xlsx", ".xls"}:
        rows = _load_excel(source)
    else:
        raise ValueError(f"Unsupported input format: {source.suffix}")

    problems = [_normalize_problem_row(row, idx) for idx, row in enumerate(rows, start=1)]
    return problems


def _load_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, start=1):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            if not isinstance(obj, dict):
                raise ValueError(f"JSONL line {line_no} is not an object")
            rows.append(obj)
    return rows


def _load_json(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        for key in ("problems", "data", "items", "questions"):
            if isinstance(data.get(key), list):
                rows = data[key]
                break
        else:
            rows = [data]
    else:
        raise ValueError("JSON input must be an object or array")
    if not all(isinstance(row, dict) for row in rows):
        raise ValueError("JSON problem rows must be objects")
    return list(rows)


def _load_csv(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _load_excel(path: Path) -> List[Dict[str, Any]]:
    try:
        import pandas as pd  # type: ignore

        return pd.read_excel(path).fillna("").to_dict(orient="records")
    except ImportError:
        pass

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ImportError("Reading Excel requires pandas or openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    data = []
    for row in rows[1:]:
        data.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
    return data


def _normalize_problem_row(row: Dict[str, Any], index: int) -> Dict[str, Any]:
    pid = _first_present(row, ID_FIELDS)
    text = _first_present(row, TEXT_FIELDS)
    if pid is None:
        pid = f"{index:03d}"
    if text is None:
        text = ""

    pid_str = str(pid).strip() or f"{index:03d}"
    text_str = str(text).strip()
    metadata = {key: value for key, value in row.items() if key not in ID_FIELDS + TEXT_FIELDS}
    return {
        "problem_id": pid_str,
        "problem_text": text_str,
        "raw_metadata": metadata,
    }


def _first_present(row: Dict[str, Any], fields: Iterable[str]) -> Optional[Any]:
    lower_map = {str(key).lower(): key for key in row.keys()}
    for field in fields:
        if field in row and row[field] not in (None, ""):
            return row[field]
        key = lower_map.get(field.lower())
        if key is not None and row[key] not in (None, ""):
            return row[key]
    return None


def _read_existing_results(path: Path) -> Dict[str, Dict[str, Any]]:
    results: Dict[str, Dict[str, Any]] = {}
    if not path.exists():
        return results
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                pid = str(obj.get("problem_id", "")).strip()
                if pid:
                    results[pid] = obj
            except Exception:
                continue
    return results


def _load_and_validate_results(path: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    results: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                pid = str(obj.get("problem_id") or f"line_{line_no}")
                solution = parse_and_validate(json.dumps(obj, ensure_ascii=False), pid)
                results.append(solution.model_dump(mode="json"))
            except Exception as exc:
                errors.append({"line": line_no, "error": f"{type(exc).__name__}: {exc}"})
    return results, errors


def _write_problem_log(log_dir: Path, problem_id: str, run_log: Dict[str, Any]) -> None:
    safe_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(problem_id)) or "unknown"
    path = log_dir / f"{safe_id}.json"
    path.write_text(json.dumps(run_log, ensure_ascii=False, indent=2), encoding="utf-8")


def _resolve_api_config(args: argparse.Namespace) -> Tuple[Optional[str], Optional[str]]:
    api_key = args.api_key or os.environ.get("OPENAI_API_KEY")
    api_base = args.api_base or os.environ.get("LLM_API_BASE")
    return api_key, api_base


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="MathSolve-Agent single-agent math solver")
    parser.add_argument("--input", "-i", type=str, default=None, help="Input JSON/JSONL/CSV/XLSX")
    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default="outputs/results.jsonl",
        help="Incremental JSONL output path",
    )
    parser.add_argument("--results-json", type=str, default=None, help="Merged JSON array output")
    parser.add_argument("--log-dir", type=str, default=None, help="Per-problem log directory")
    parser.add_argument("--summary", type=str, default=None, help="Run summary JSON path")
    parser.add_argument("--model", "-m", type=str, default="gpt-4o-mini", help="Model name")
    parser.add_argument("--api-key", type=str, default=None, help="API key")
    parser.add_argument("--api-base", type=str, default=None, help="OpenAI-compatible chat endpoint")
    parser.add_argument("--limit", "-n", type=int, default=None, help="Only process first N rows")
    parser.add_argument("--resume", action="store_true", help="Skip IDs already in output JSONL")
    parser.add_argument("--demo", action="store_true", help="Run a single demo problem")
    parser.add_argument("--config", type=str, default=None, help="JSON/YAML runtime config")
    parser.add_argument(
        "--official",
        action="store_true",
        help="Fail fast unless Intern-S1 model, token, and InternLM chat endpoint are configured",
    )
    parser.add_argument(
        "--ablation",
        type=str,
        default="full",
        help=(
            "Ablation preset: full, official_stable, strong, no_sandbox, "
            "no_ortools, no_normalizer, no_equivalence, strict_equivalence, "
            "no_llm_verify, no_extract, single_candidate"
        ),
    )
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    api_key, api_base = _resolve_api_config(args)

    if args.demo or not args.input:
        run_single_demo(
            args.model,
            api_key,
            api_base,
            args.config,
            args.ablation,
            args.official,
        )
        return

    run_batch(
        input_path=args.input,
        output_path=args.output,
        model_type=args.model,
        api_key=api_key,
        api_base=api_base,
        limit=args.limit,
        resume=args.resume,
        results_json_path=args.results_json,
        log_dir=args.log_dir,
        summary_path=args.summary,
        config_path=args.config,
        ablation=args.ablation,
        official_mode=args.official,
    )


if __name__ == "__main__":
    main()
